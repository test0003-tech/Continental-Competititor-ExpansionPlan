import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load all clean brand data
brands = {}
for brand in ['bridgestone', 'apollo', 'yokohama', 'michelin', 'continental', 'goodyear', 'mrf']:
    fname = f'/home/z/my-project/download/{brand}_dealers_clean.json'
    with open(fname) as f:
        brands[brand] = json.load(f)

# Create workbook
wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
brand_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
data_font = Font(name='Calibri', size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Column headers
columns = ['Brand', 'Dealer Name', 'Address', 'City', 'State', 'Pincode', 'Phone', 'Email', 'Latitude', 'Longitude', 'Dealer Type']

# ---- Sheet 1: All Brands Combined ----
ws = wb.active
ws.title = 'All Brands'

# Write headers
for col, header in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Write data
row = 2
brand_colors = {
    'bridgestone': 'E2EFDA',
    'apollo': 'FCE4D6',
    'yokohama': 'D9E2F3',
    'michelin': 'FFF2CC',
    'continental': 'E2D9F3',
    'goodyear': 'D6E4F0',
    'mrf': 'F8D7DA',
}

total_dealers = 0
total_with_coords = 0

for brand_name, data in brands.items():
    brand_color = brand_colors.get(brand_name, 'FFFFFF')
    fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
    
    for d in data:
        brand_display = brand_name.upper()
        values = [
            brand_display,
            d.get('name', ''),
            d.get('address', ''),
            d.get('city', ''),
            d.get('state', ''),
            d.get('pincode', ''),
            d.get('phone', ''),
            d.get('email', ''),
            d.get('latitude', ''),
            d.get('longitude', ''),
            d.get('dealer_type', ''),
        ]
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(val) if val else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = wrap_alignment
            cell.fill = fill
        
        total_dealers += 1
        if d.get('latitude') and d.get('longitude'):
            total_with_coords += 1
        
        row += 1

# Set column widths
col_widths = [12, 35, 40, 18, 20, 10, 18, 25, 14, 14, 15]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze top row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:K{row-1}'

print(f"All Brands sheet: {total_dealers} dealers, {total_with_coords} with real lat/lng")

# ---- Individual brand sheets ----
for brand_name, data in brands.items():
    ws = wb.create_sheet(title=brand_name.upper())
    
    # Write headers
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Write data
    for r, d in enumerate(data, 2):
        values = [
            brand_name.upper(),
            d.get('name', ''),
            d.get('address', ''),
            d.get('city', ''),
            d.get('state', ''),
            d.get('pincode', ''),
            d.get('phone', ''),
            d.get('email', ''),
            d.get('latitude', ''),
            d.get('longitude', ''),
            d.get('dealer_type', ''),
        ]
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=str(val) if val else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = wrap_alignment
    
    # Set column widths
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = f'A1:K{len(data)+1}'
    
    has_coords = sum(1 for d in data if d.get('latitude') and d.get('longitude'))
    print(f"{brand_name.upper()} sheet: {len(data)} dealers, {has_coords} with real lat/lng")

# ---- Summary sheet ----
ws = wb.create_sheet(title='SUMMARY', index=0)

# Summary headers
summary_headers = ['Brand', 'Total Dealers', 'With Real Lat/Lng', 'Without Lat/Lng', 'Coverage %']
for col, header in enumerate(summary_headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Summary data
row = 2
grand_total = 0
grand_coords = 0
for brand_name, data in brands.items():
    total = len(data)
    with_coords = sum(1 for d in data if d.get('latitude') and d.get('longitude'))
    without = total - with_coords
    pct = f"{with_coords*100/total:.1f}%" if total else "0%"
    
    values = [brand_name.upper(), total, with_coords, without, pct]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    grand_total += total
    grand_coords += with_coords
    row += 1

# Grand total row
grand_without = grand_total - grand_coords
grand_pct = f"{grand_coords*100/grand_total:.1f}%"
values = ['TOTAL', grand_total, grand_coords, grand_without, grand_pct]
bold_font = Font(name='Calibri', bold=True, size=10)
for col, val in enumerate(values, 1):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

# Set summary column widths
for i, w in enumerate([15, 15, 18, 18, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Note about CEAT and JK Tyre
row += 2
note_cell = ws.cell(row=row, column=1, value='Note:')
note_cell.font = Font(name='Calibri', bold=True, size=10)
row += 1
note_cell = ws.cell(row=row, column=1, value='CEAT and JK Tyre could not be scraped due to WAF/bot protection (Netacea and Cloudflare respectively).')
note_cell.font = Font(name='Calibri', italic=True, size=10, color='FF0000')
row += 1
note_cell = ws.cell(row=row, column=1, value='All latitude/longitude values are REAL coordinates from source APIs only. No geocoded data is included.')
note_cell.font = Font(name='Calibri', italic=True, size=10, color='006600')

# Save
output_path = '/home/z/my-project/download/India_Tyre_Dealers_All_Brands.xlsx'
wb.save(output_path)
print(f"\nExcel saved to: {output_path}")
print(f"Total: {grand_total} dealers, {grand_coords} with real lat/lng ({grand_pct})")
